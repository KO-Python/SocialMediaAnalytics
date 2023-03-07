import json
import time
import requests # requests 모듈 설치 필요!
import openpyxl
import os
from chromedriver_autoinstaller import install
from selenium import webdriver
import datetime

from selenium.webdriver.common.by import By

url = input("유튜브 URL 주소 입력 >> ")

print("[알림] 유튜브 댓글을 로딩 중입니다. 최대 1분 소요 (평균 10초 소요)")
opt = webdriver.ChromeOptions()
opt.add_argument("headless")
browser = webdriver.Chrome(install(), options=opt)
browser.implicitly_wait(10)
browser.get('https://hadzy.com/')
browser.find_element(By.CSS_SELECTOR, "button.MuiButtonBase-root.MuiButton-root.MuiButton-text.sc-gqjmRU.fcuJZS").click()
browser.find_element(By.CSS_SELECTOR, "input.MuiInputBase-input").send_keys(url)
browser.find_element(By.CSS_SELECTOR, "button.MuiButtonBase-root.MuiIconButton-root.sc-iAyFgw.cZnNCh").click()
time.sleep(3)
browser.find_element(By.CSS_SELECTOR, 'button.MuiButtonBase-root.MuiFab-root').click()
time.sleep(3)
while True:
    temp = browser.find_element(By.CSS_SELECTOR, "button.MuiButtonBase-root.MuiFab-root").text
    if temp == "View Comments":
        break
browser.close()

url_v = url.replace("https://www.youtube.com/watch?v=", "")
h = {
    'Referer': 'https://hadzy.com/',
    'sec-ch-ua': '"Whale";v="3", " Not;A Brand";v="99", "Chromium";v="96"',
    'sec-ch-ua-mobile': '?0',
    'sec-ch-ua-platform': 'macOS',
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.110 Whale/3.12.129.46 Safari/537.36'
}

if not os.path.exists("./유튜브댓글수집"):
    os.mkdir("./유튜브댓글수집")

today = datetime.datetime.today().strftime("%m월%d일_%H시%M분%S초")
file_name = f"./유튜브댓글수집/유튜브댓글수집_{today}.xlsx"
# 엑셀 파일 생성
if not os.path.exists(file_name):
    openpyxl.Workbook().save(file_name)

try:
    book = openpyxl.load_workbook(file_name)
except:
    print("[에러] 기존에 저장돼있는 엑셀파일을 삭제해주세요.")
# 쓸데 없는 시트는 삭제하기
if "Sheet" in book.sheetnames:
    book.remove(book["Sheet"])
sheet = book.create_sheet()
sheet.title = "YouTube comments"

page_num = 0
row_num = 1
print("[알림] 이제 댓글 크롤링을 시작합니다.")
print("--------------------------------------------")
while True:
    p = {
        'page': f'{page_num}',
        'size': '10',
        'sortBy': 'publishedAt',
        'direction': 'asc',
        'searchTerms':'',
        'author':''
    }
    res = requests.get(f"https://hadzy.com/api/comments/{url_v}", params=p, headers=h)
    # print(res.text)
    content_list = json.loads(res.text)['content']
    if len(content_list) == 0:
        print("--------------------------------------------")
        print("[알림] 크롤링이 끝났습니다.")
        break
    for i in content_list:
        comment = i['textDisplay'].replace("<br>", " ")
        print(comment)
        # 엑셀에 기록
        sheet.cell(row=row_num, column=1).value = comment
        book.save(file_name)
        row_num+=1
    page_num += 1

