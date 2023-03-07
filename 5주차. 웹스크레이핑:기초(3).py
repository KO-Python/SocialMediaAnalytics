
"""url구조파악
https://search.naver.com/search.naver?                   # 기본 url(네이버 검색)
where=news                                               # 기사 검색 (블로그나 카페 아닌 기사)
&query=%EA%B5%90%ED%86%B5%EC%82%AC%EA%B3%A0              # 검색어('교통사고')
&sort=0                                                  # 정렬   0=관련도순, 1=최신순, 2=오래된순
&photo=0                                                 # 포토   0=전체, 1=동영상, 2=포토, 3=지면기사, 4=보도자료, 5=자동생성기사
&field=1                                                 # 영역   0=전체, 1= 제목                       
&nso=so%3Ar%2Cp%3Afrom20210101to20210228%2Ca%3Aall       # 시작 날짜 ~ 마지막 날짜
&start=1                                                 # 페이지당 시작 기사 번호로 1, 11, 21, 31 등으로 가능. 1페이지는 1부터, 2페이지는 11부터, 3페이지는 21부터.
"""







#방법1: 네이터 뉴스에서 특정 검색어 입력 후 가장 상위에 있는 뉴스와 링크 스크레이핑
import requests
from bs4 import BeautifulSoup
keyword = '코로나'
r = requests.get(f'https://search.naver.com/search.naver?where=news&query={keyword}')
soup = BeautifulSoup(r.text, 'html.parser')
articles = soup.select('ul.list_news > li')
title = articles[0].select_one('a.news_tit')['title']
url = articles[0].select_one('div.info_group > a:nth-of-type(2)')['href']
print('title=', title)
print('url=', url)








#방법2: 네이터 뉴스에서 특정 검색어 입력 후 제목과 출처
keyword = '윤석열'
r = requests.get(f"https://search.naver.com/search.naver?where=news&query={keyword}",
                 headers={'User-Agent':'Mozilla/5.0'})
html = BeautifulSoup(r.text, "html.parser")
articles = html.select("ul.list_news > li")
title = articles[0].select_one("a.news_tit").text
source = articles[0].select_one("a.info.press").text
print(title, source)







#방법3: 수집반복추출
keyword = '윤석열'
r = requests.get(f"https://search.naver.com/search.naver?where=news&query={keyword}",
                 headers={'User-Agent':'Mozilla/5.0'})
html = BeautifulSoup(r.text, "html.parser")
articles = html.select("ul.list_news > li")
for r in articles:
    title = r.select_one("a.news_tit").text
    source = r.select_one("a.info.press").text
    print(title, source)









"""
환율정보 가져오기
"""
import pandas as pd
url = "https://search.naver.com/search.naver?where=nexearch&sm=top_hty&fbm=1&ie=utf8&query=%ED%99%98%EC%9C%A8"
dfs = pd.read_html(url)
dfs
len(dfs) #데이터프레임 내 요소 개수 확인하기
dfs[0] #데이터 확인
exchange_rate_df = dfs[0].replace({'전일대비상승': '▲','전일대비하락': '▼'}, regex=True) #문자열 변환
exchange_rate_df








#네이버부동산동향보고서 가져오기
import pandas as pd
base_url = 'https://land.naver.com/news/trendReport.naver'
page_num = 1
url = f'{base_url}?page={page_num}'
dfs = pd.read_html(url)
df = dfs[0]
df
# 행과 열의 최대 표시 개수를 임시로 설정
with pd.option_context('display.max_rows',4, 'display.max_columns',6):
    pd.set_option("show_dimensions", False) # 행과 열 개수 정보 숨기기
# 원본 DataFrame의 제목 열에 있는 문자열을 분리해
# 전국, 서울, 수도권의 매매가 변화율 열이 있는 DataFrame 반환하는 함수
def split_title_to_rates(df_org):
    df_new = df_org.copy()
    df_temp = df_new['제목'].str.replace('%','')  # 제목 문자열에서 % 제거
    df_temp = df_temp.str.replace('보합', '0')  # 제목 문자열에서 보합을 0으로 바꿈
    df_temp = df_temp.str.replace('보합세', '0')  # 제목 문자열에서 보합세를 0으로 바꿈
    regions = ['전국', '서울', '수도권']
    for region in regions:
        df_temp = df_temp.str.replace(region, '')  # 문자열에서 전국, 서울, 수도권 제거
    df_temp = df_temp.str.split(']', expand=True)  # ]를 기준으로 열 분리
    df_temp = df_temp[1].str.split(',', expand=True)  # ,를 기준으로 열 분리
    df_temp = df_temp.astype(float)
    df_new[regions] = df_temp  # 전국, 서울, 수도권 순서대로 DataFrame 데이터에 할당
    return df_new[['등록일'] + regions + ['번호']]  # DataFrame에서 필요한 열만 반환
df_rate = split_title_to_rates(df) # split_title_to_rates() 함수 호출
df_rate.head()                     # 앞의 일부만 출력
#위 코드를 활용해 동향보고서 1부터 8페이지 제목을 가져와 전국, 서울, 수도원 아파트 매매가 변화율 추출
base_url = "https://land.naver.com/news/trendReport.naver"
df_rates = pd.DataFrame()  # 전체 데이터가 담길 DataFrame 데이터
last_page_num = 8  # 가져올 데이터의 마지막 페이지
for page_num in range(1, last_page_num + 1):
    url = f"{base_url}?page={page_num}"
    dfs = pd.read_html(url)
    df_page = dfs[0]  # 리스트의 첫 번째 항목에 동향 보고서 제목 데이터가 있음
    df_rate = split_title_to_rates(df_page)
    # 세로 방향으로 연결 (기존 index를 무시)
    df_rates = pd.concat([df_rates, df_rate], ignore_index=True)
# 최신 데이터와 과거 데이터의 순서를 바꿈. index도 초기화함
df_rates = df_rates[::-1].reset_index(drop=True)
df_rates.head()  # 앞의 일부만 출력성
# 행과 열의 최대 표시 개수를 임시로 설정
with pd.option_context('display.max_rows', 4, 'display.max_columns', 6):
    pd.set_option("show_dimensions", False)  # 행과 열 개수 정보 숨기기
import matplotlib as mpl
mpl.rcParams['font.family'] = 'AppleGothic' # '맑은 고딕'으로 폰트 설정
mpl.rcParams['axes.unicode_minus'] = False # 마이너스(-) 폰트 깨짐 방지
import pandas as pd
import matplotlib.pyplot as plt
df_rates.head(40).plot(x='등록일', y=['전국', '서울', '수도권'], #최신 10개 데이터 이용 그래프 생성
                       figsize=(10, 8), subplots=True, layout=(3,1),
                       style = '-o', grid=True) # 그래프 그리기







"""
연습#1:뉴스추출 
"""
#중앙일보 스크레이핑
from bs4 import BeautifulSoup
import urllib.request as req
import urllib.parse as par

keyword = input("키워드 입력 >> ")
encoded = par.quote(keyword)  # 한글 -> 특수한 문자

page_num = 1
while True:
    url = f"https://www.joongang.co.kr/_CP/496?keyword={encoded}&sort%20=&pageItemId=439&page={page_num}"
    code = req.urlopen(url)
    soup = BeautifulSoup(code, "html.parser")
    title = soup.select("h2.headline a")
    if len(title) == 0:  # 끝 페이지까지 크롤링 완료했으면?
        break
    for i in title:
        print("제목 :", i.text.strip())
        print("링크 :", i.attrs["href"])
        code_news = req.urlopen(i.attrs["href"])
        soup_news = BeautifulSoup(code_news, "html.parser")
        content = soup_news.select_one("div#article_body")
        result = content.text.strip().replace("     ", " ").replace("   ", "")
        print(result)
        print()
    page_num += 1








"""
#연습2: 날씨 가져오기  
"""

import requests
from bs4 import BeautifulSoup
import time


def get_weather_daum(location):
    search_query = location + " 날씨"
    search_url = "https://search.daum.net/search?w=tot&DA=YZR&t__nil_searchbox=btn&sug=&sugo=&sq=&o=&q="
    url = search_url + search_query
    html_weather = requests.get(url).text
    time.sleep(2)
    soup_weather = BeautifulSoup(html_weather, "lxml")

    txt_temp = soup_weather.select_one('strong.txt_temp').get_text()
    txt_weather = soup_weather.select_one('span.txt_weather').get_text()

    dl_weather_dds = soup_weather.select('dl.dl_weather dd')
    [wind_speed, humidity, pm10] = [x.get_text() for x in dl_weather_dds]

    return (txt_temp, txt_weather, wind_speed, humidity, pm10)
location = "서울시 종로구 청운동"
get_weather_daum(location)








"""
#연습3: 웹 크롤링 사례 
"""
import urllib.request as req
from bs4 import BeautifulSoup
import os
import openpyxl
import datetime
from openpyxl.drawing.image import Image
import re # 추가






# 이미지 저장할 폴더 생성
if not os.path.exists("./멜론이미지"):
    os.mkdir("./멜론이미지")

header = req.Request("https://www.melon.com/chart/index.htm", headers={"User-Agent":"Mozilla/5.0"})
code = req.urlopen(header)
soup = BeautifulSoup(code, "html.parser")
title = soup.select("div.ellipsis.rank01 > span > a")
name = soup.select("div.ellipsis.rank02 > span")
album = soup.select("div.ellipsis.rank03 > a")
img = soup.select("a.image_typeAll > img")






# 엑셀 파일 생성
if not os.path.exists("./멜론_크롤링.xlsx"):
    openpyxl.Workbook().save("./멜론_크롤링.xlsx")

book = openpyxl.load_workbook("./멜론_크롤링.xlsx")

# 쓸데 없는 시트는 삭제하기
if "Sheet" in book.sheetnames:
    book.remove(book["Sheet"])
sheet = book.create_sheet()
now = datetime.datetime.now()
sheet.title = "{}년 {}월 {}일 {}시 {}분 {}초".format(now.year, now.month, now.day, now.hour, now.minute, now.second)
row_num = 1






# 열 너비 조절
sheet.column_dimensions["A"].width = 15
sheet.column_dimensions["B"].width = 50
sheet.column_dimensions["C"].width = 29
sheet.column_dimensions["D"].width = 47
for i in range(len(title)):
    img_file_name = "./멜론이미지/{}.png".format(re.sub("[\\\/:*?\"<>\|]", " ", album[i].string))  # 수정
    req.urlretrieve(img[i].attrs["src"], img_file_name) # 수정
    # 엑셀에 크롤링 결과 출력
    img_for_excel = Image(img_file_name)
    sheet.add_image(img_for_excel, "A{}".format(row_num))
    sheet.cell(row=row_num, column=2).value = title[i].string
    sheet.cell(row=row_num, column=3).value = name[i].text
    sheet.cell(row=row_num, column=4).value = album[i].string
    sheet.row_dimensions[row_num].height = 90 # 행높이 조절.
    book.save("./멜론_크롤링.xlsx")
    print("{}위. {} - {}".format(row_num, title[i].string, name[i].text))
    row_num+=1
